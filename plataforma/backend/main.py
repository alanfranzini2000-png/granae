import os
import io
import json
import random
from concurrent.futures import ThreadPoolExecutor
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Header, Query, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta

from database import (
    init_db, get_conn, atualizar_status_base,
    set_usuario, get_usuario, sanitizar_usuario, USUARIO_PADRAO,
    listar_usuarios, criar_usuario, apagar_usuario, zerar_usuario,
    buscar_apelido,
)
from categorizer import (
    categorizar, categorizar_lote, categorizar_por_regra, limpar_desc, eh_fatura,
    eh_imune_viagem, ANTHROPIC_API_KEY, CATEGORIAS_VALIDAS, KEYWORDS, ORDEM,
)
from parser import processar_pdf, pdf_esta_encriptado
from planilha import ler_previa, categorias_distintas, parse_planilha

# Resolve o usuário (base) da requisição: header 'X-Usuario' ou query '?usuario='.
# Setado como dependency global → todo get_conn() na requisição usa a base certa.
# (EventSource e links <a> não mandam header, por isso aceitamos a query também.)
_bases_inicializadas = set()

async def resolver_usuario(
    request: Request,
    x_usuario: Optional[str] = Header(default=None),
    usuario: Optional[str] = Query(default=None),
):
    # IMPORTANTE: precisa ser async. Dependency síncrona roda em outra thread e
    # o ContextVar setado não chegaria ao endpoint (vazaria pra base errada).
    u = sanitizar_usuario(x_usuario or usuario or USUARIO_PADRAO)
    set_usuario(u)
    if u not in _bases_inicializadas:
        init_db(u)
        _bases_inicializadas.add(u)
    # DEBUG temporário: qual base cada requisição usou (header/query crus)
    print(f"[BASE] {request.method} {request.url.path} -> '{u}' (X-Usuario={x_usuario!r} q={usuario!r})", flush=True)
    return u

app = FastAPI(title="Finanças Pessoais API", dependencies=[Depends(resolver_usuario)])
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
init_db(USUARIO_PADRAO)


# ── MODELS ────────────────────────────────────────────────────────────────

class IncorporarItem(BaseModel):
    id: int
    categoria: Optional[str] = None
    excluir: Optional[bool] = False
    descricao: Optional[str] = None   # renome "só este" feito na revisão

class IncorporarBody(BaseModel):
    upload_id: int
    itens: List[IncorporarItem]

class NovaViagem(BaseModel):
    destino: str
    data_inicio: str
    data_fim: str
    card: Optional[str] = None   # nome do arquivo de imagem escolhido (galeria)

class AtualizarLancamento(BaseModel):
    categoria: Optional[str] = None
    viagem: Optional[str] = None
    revisado: Optional[int] = None
    data: Optional[str] = None
    mes: Optional[str] = None
    valor: Optional[float] = None
    descricao: Optional[str] = None

class NovaRegra(BaseModel):
    palavra_chave: str
    categoria: str

class NovaMeta(BaseModel):
    tipo: str
    categoria: Optional[str] = None
    valor_alvo: float
    reducao_modo: Optional[str] = None
    acumulo_meses: Optional[int] = None
    acumulo_mes_fim: Optional[str] = None


# ── HELPERS ───────────────────────────────────────────────────────────────

def _regras_usuario(conn):
    """Mapa mental do perfil ativo: lista de (palavra_chave, categoria)."""
    rows = conn.execute(
        "SELECT palavra_chave, categoria FROM regras_categorias ORDER BY id"
    ).fetchall()
    return [(r['palavra_chave'], r['categoria']) for r in rows]


def _processar_arquivo(conteudo, nome_arquivo, conn, upload_id, senha=None):
    tipo, lancamentos_raw = processar_pdf(conteudo, nome_arquivo, senha)
    regras_usuario = _regras_usuario(conn)

    # 1) Pré-processa SEM IA: limpa a descrição, filtra faturas (débito), apelido.
    itens = []
    for l in lancamentos_raw:
        is_pix = 'PIX ENVIADO' in l['descricao'].upper() or 'PIX RECEBIDO' in l['descricao'].upper()
        desc_real = limpar_desc(l['descricao'])
        if tipo == 'debito' and eh_fatura(desc_real):
            continue
        apelido = buscar_apelido(conn, desc_real)   # nome fantasia (se houver)
        itens.append({'l': l, 'is_pix': is_pix, 'desc_real': desc_real, 'desc': apelido or desc_real})

    # 2) Categoriza com DEDUPLICAÇÃO + chamada em LOTE. Cada descrição única vai
    #    1x à IA (não uma por ocorrência) e TODAS as descrições sem regra vão numa
    #    ÚNICA chamada à IA — o que era O(n) chamadas sequenciais vira ~1. Chave
    #    inclui o "PIX pequeno (≤100)" porque essa faixa muda a categoria de PIX.
    def _chave(it):
        pix_pequeno = it['is_pix'] and abs(it['l']['valor']) <= 100
        return (it['desc_real'].upper(), it['l']['tipo'], it['is_pix'], pix_pequeno)

    unicos = {}
    for it in itens:
        unicos.setdefault(_chave(it), it)

    lote = [{'descricao': it['desc_real'], 'valor': it['l']['valor'],
             'tipo': it['l']['tipo'], 'is_pix': it['is_pix']} for it in unicos.values()]
    resultados = categorizar_lote(lote, regras_usuario=regras_usuario)
    cache = {k: res for k, res in zip(unicos.keys(), resultados)}

    # 3) Monta o staging usando o resultado em cache.
    staging = []
    for it in itens:
        l = it['l']; desc = it['desc']; desc_real = it['desc_real']; valor = l['valor']
        cat, confianca, fonte = cache[_chave(it)]
        credito = valor if valor > 0 else None
        debito  = abs(valor) if valor < 0 else None

        dup = conn.execute("""
            SELECT COUNT(*) as n FROM lancamentos
            WHERE data=? AND descricao=? AND valor=?
        """, (l['data'], desc, valor)).fetchone()['n']

        cur = conn.execute("""
            INSERT INTO staging
              (upload_id, mes, data, descricao, descricao_real, credito, debito, valor,
               categoria, tipo, fonte, confianca, duplicata)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (upload_id, l['mes'], l['data'], desc, desc_real,
              credito, debito, valor,
              cat, l['tipo'], fonte or 'desconhecido', confianca or 'vermelho', 1 if dup > 0 else 0))

        staging.append({
            'id':        cur.lastrowid,
            'mes':       l['mes'],
            'data':      l['data'],
            'descricao': desc,
            'descricao_real': desc_real,
            'valor':     valor,
            'categoria': cat,
            'tipo':      l['tipo'],
            'confianca': confianca or 'vermelho',
            'fonte':     fonte or 'desconhecido',
            'duplicata': dup > 0,
            'arquivo':   tipo,
        })
    return tipo, staging


# ── UPLOAD → STAGING (múltiplos arquivos) ─────────────────────────────────

@app.post("/upload")
async def upload_arquivos(
    files: List[UploadFile] = File(...),
    senhas: str = Form(default="{}")
):
    """
    Aceita 1 ou mais PDFs. Detecta tipo automaticamente para cada um.
    Duplicatas são removidas dos itens retornados mas contabilizadas.
    PDFs com senha: retorna arquivos_com_senha para o frontend pedir a senha.
    """
    try:
        senhas_dict = json.loads(senhas)
    except Exception:
        senhas_dict = {}

    # Lê todos os arquivos primeiro para poder verificar criptografia
    arquivos_dados = []
    for file in files:
        conteudo = await file.read()
        arquivos_dados.append((file.filename, conteudo))

    # Verifica quais PDFs estão protegidos e sem senha fornecida
    precisam_senha = [
        nome for nome, conteudo in arquivos_dados
        if pdf_esta_encriptado(conteudo) and nome not in senhas_dict
    ]
    if precisam_senha:
        return {"arquivos_com_senha": precisam_senha}

    conn = get_conn()

    cur = conn.execute("""
        INSERT INTO uploads (tipo, nome_arquivo, periodo_inicio, periodo_fim, total_lancamentos, incorporado)
        VALUES (?,?,?,?,?,0)
    """, ('multiplo', ' + '.join(n for n, _ in arquivos_dados), None, None, 0))
    upload_id = cur.lastrowid

    todos_staging = []
    erros = []

    for nome, conteudo in arquivos_dados:
        senha = senhas_dict.get(nome)
        try:
            tipo, staging = _processar_arquivo(conteudo, nome, conn, upload_id, senha)
            todos_staging.extend(staging)
        except Exception as e:
            erros.append(f"{nome}: {str(e)}")

    # Separar duplicatas dos itens a mostrar
    itens_limpos    = [s for s in todos_staging if not s['duplicata']]
    total_duplicatas = sum(1 for s in todos_staging if s['duplicata'])

    # Remover do staging as duplicatas (não precisam de revisão)
    ids_dup = [s['id'] for s in todos_staging if s['duplicata']]
    if ids_dup:
        conn.execute(f"DELETE FROM staging WHERE id IN ({','.join('?'*len(ids_dup))})", ids_dup)

    total = len(itens_limpos)
    conn.execute("UPDATE uploads SET total_lancamentos=? WHERE id=?", (total, upload_id))
    conn.commit()
    conn.close()

    return {
        "upload_id":        upload_id,
        "total":            total,
        "duplicatas_removidas": total_duplicatas,
        "sem_categoria":    sum(1 for s in itens_limpos if not s['categoria']),
        "itens":            itens_limpos,
        "erros":            erros,
    }


# ── IMPORTAR PLANILHA (base antiga via Excel) → STAGING ───────────────────

@app.post("/importar-planilha/previa")
async def importar_planilha_previa(file: UploadFile = File(...)):
    """Lê as primeiras linhas do Excel para o usuário apontar linha/colunas."""
    conteudo = await file.read()
    try:
        return ler_previa(conteudo)
    except Exception as e:
        raise HTTPException(400, f"Não foi possível ler a planilha: {e}")


@app.post("/importar-planilha/categorias")
async def importar_planilha_categorias(
    file: UploadFile = File(...),
    col_categoria: int = Form(...),
    linha_inicio: int = Form(...),
):
    """Valores distintos da coluna de categoria + de-para sugerido (heurística)."""
    conteudo = await file.read()
    try:
        return {"categorias": categorias_distintas(conteudo, col_categoria, linha_inicio)}
    except Exception as e:
        raise HTTPException(400, f"Erro ao analisar as categorias: {e}")


@app.post("/importar-planilha")
async def importar_planilha(
    file: UploadFile = File(...),
    mapeamento: str = Form(...),
    de_para: str = Form(default="{}"),
):
    """Lê a planilha com o mapeamento + de-para e insere no staging (vai p/ Revisão)."""
    conteudo = await file.read()
    try:
        mp = json.loads(mapeamento)
        dp = json.loads(de_para)
    except Exception:
        raise HTTPException(400, "Mapeamento ou de-para inválido.")

    for campo in ('col_data', 'col_descricao', 'col_valor', 'linha_inicio'):
        if mp.get(campo) is None:
            raise HTTPException(400, f"Mapeamento incompleto: falta '{campo}'.")

    try:
        lancs = parse_planilha(conteudo, mp, dp)
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler a planilha: {e}")
    if not lancs:
        raise HTTPException(400, "Nenhuma linha válida encontrada com esse mapeamento. "
                                 "Confira a linha do 1º dado e as colunas de data/valor.")

    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO uploads (tipo, nome_arquivo, periodo_inicio, periodo_fim, total_lancamentos, incorporado)
        VALUES (?,?,?,?,?,0)
    """, ('planilha', file.filename or 'planilha.xlsx', None, None, 0))
    upload_id = cur.lastrowid

    todos = []
    for l in lancs:
        valor = l['valor']
        credito = valor if valor > 0 else None
        debito  = abs(valor) if valor < 0 else None
        arquivo = 'credito' if valor >= 0 else 'debito'
        conf = 'verde' if l['categoria'] else 'vermelho'
        dup = conn.execute("""
            SELECT COUNT(*) as n FROM lancamentos
            WHERE data=? AND descricao=? AND valor=?
        """, (l['data'], l['descricao'], valor)).fetchone()['n']

        cur2 = conn.execute("""
            INSERT INTO staging
              (upload_id, mes, data, descricao, credito, debito, valor,
               categoria, tipo, fonte, confianca, duplicata)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (upload_id, l['mes'], l['data'], l['descricao'],
              credito, debito, valor,
              l['categoria'], l['tipo'], 'importacao', conf, 1 if dup > 0 else 0))

        todos.append({
            'id':        cur2.lastrowid,
            'mes':       l['mes'],
            'data':      l['data'],
            'descricao': l['descricao'],
            'valor':     valor,
            'categoria': l['categoria'],
            'tipo':      l['tipo'],
            'confianca': conf,
            'fonte':     'importacao',
            'duplicata': dup > 0,
            'arquivo':   arquivo,
        })

    itens_limpos = [s for s in todos if not s['duplicata']]
    total_dup    = sum(1 for s in todos if s['duplicata'])
    ids_dup      = [s['id'] for s in todos if s['duplicata']]
    if ids_dup:
        conn.execute(f"DELETE FROM staging WHERE id IN ({','.join('?'*len(ids_dup))})", ids_dup)

    total = len(itens_limpos)
    conn.execute("UPDATE uploads SET total_lancamentos=? WHERE id=?", (total, upload_id))
    conn.commit()
    conn.close()

    return {
        "upload_id":            upload_id,
        "total":                total,
        "duplicatas_removidas": total_dup,
        "sem_categoria":        sum(1 for s in itens_limpos if not s['categoria']),
        "itens":                itens_limpos,
        "erros":                [],
    }


# ── INCORPORAR ────────────────────────────────────────────────────────────

@app.post("/incorporar")
async def incorporar(body: IncorporarBody):
    conn = get_conn()
    incorporados = 0
    ignorados_duplicata = 0
    tipos_incorporados = set()

    for item in body.itens:
        if item.excluir or not item.categoria:
            continue
        row = conn.execute("SELECT * FROM staging WHERE id=?", (item.id,)).fetchone()
        if not row:
            continue
        # Descrição final: override do usuário (renome "só este") ou a do staging.
        desc_final = item.descricao if item.descricao else row['descricao']
        desc_real  = row['descricao_real'] if 'descricao_real' in row.keys() else None
        ja_existe = conn.execute("""
            SELECT COUNT(*) as n FROM lancamentos
            WHERE data=? AND descricao=? AND valor=?
        """, (row['data'], desc_final, row['valor'])).fetchone()['n']
        if ja_existe:
            ignorados_duplicata += 1
            continue
        conn.execute("""
            INSERT INTO lancamentos
              (mes, data, descricao, descricao_real, credito, debito, valor,
               categoria, tipo, revisado, origem, fonte, confianca)
            VALUES (?,?,?,?,?,?,?,?,?,1,'upload',?,?)
        """, (row['mes'], row['data'], desc_final, desc_real,
              row['credito'], row['debito'], row['valor'],
              item.categoria, row['tipo'], row['fonte'], row['confianca']))
        tipos_incorporados.add(row['tipo'])
        incorporados += 1

    conn.execute("DELETE FROM staging WHERE upload_id=?", (body.upload_id,))
    conn.execute("UPDATE uploads SET incorporado=1 WHERE id=?", (body.upload_id,))

    if 'Débito' in tipos_incorporados:
        atualizar_status_base(conn, 'debito')
    if 'Crédito' in tipos_incorporados:
        atualizar_status_base(conn, 'credito')

    conn.commit()
    conn.close()
    return {"ok": True, "incorporados": incorporados, "ignorados_duplicata": ignorados_duplicata}


# ── REGRAS (aprendizado) ──────────────────────────────────────────────────

@app.post("/regras")
def adicionar_regra(body: NovaRegra):
    """Salva nova regra fixa: palavra_chave → categoria."""
    conn = get_conn()
    conn.execute("""
        INSERT OR REPLACE INTO regras_categorias (palavra_chave, categoria)
        VALUES (?, ?)
    """, (body.palavra_chave.upper().strip(), body.categoria))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/regras")
def listar_regras():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM regras_categorias ORDER BY categoria, palavra_chave").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.delete("/regras/{palavra}")
def remover_regra(palavra: str):
    """Remove uma regra do mapa mental do perfil."""
    conn = get_conn()
    conn.execute("DELETE FROM regras_categorias WHERE palavra_chave=?", (palavra.upper().strip(),))
    conn.commit(); conn.close()
    return {"ok": True}


# ── MAPA MENTAL (regras do usuário + padrões do sistema) ──────────────────

@app.get("/mapa")
def mapa_mental():
    """Por categoria: padrões do sistema (só leitura) e regras do usuário (editáveis)."""
    conn = get_conn()
    user_rows = conn.execute(
        "SELECT palavra_chave, categoria FROM regras_categorias ORDER BY palavra_chave"
    ).fetchall()
    conn.close()
    por_cat = {}
    for r in user_rows:
        por_cat.setdefault(r['categoria'], []).append(r['palavra_chave'])
    categorias = [
        {"cat": cat, "padroes": KEYWORDS.get(cat, []), "usuario": por_cat.get(cat, [])}
        for cat in ORDEM
    ]
    return {"categorias": categorias}


# ── INSIGHTS IGNORADOS (suprimidos até o fim do mês) ───────────────────────

class IgnorarInsight(BaseModel):
    chave: str

def _ultimo_dia_mes_iso():
    hoje = datetime.now()
    prox = datetime(hoje.year + 1, 1, 1) if hoje.month == 12 else datetime(hoje.year, hoje.month + 1, 1)
    return (prox - timedelta(days=1)).strftime("%Y-%m-%d")

@app.post("/insights/ignorar")
def ignorar_insight(body: IgnorarInsight):
    chave = (body.chave or "").strip()
    if not chave:
        raise HTTPException(400, "chave vazia")
    ate = _ultimo_dia_mes_iso()
    conn = get_conn()
    conn.execute("""
        INSERT INTO insights_ignorados (chave, ignorado_ate) VALUES (?, ?)
        ON CONFLICT(chave) DO UPDATE SET ignorado_ate=excluded.ignorado_ate
    """, (chave, ate))
    conn.commit(); conn.close()
    return {"ok": True, "ignorado_ate": ate}

@app.get("/insights/ignorados")
def listar_insights_ignorados():
    """Chaves ainda válidas (ignorado_ate >= hoje)."""
    hoje = datetime.now().strftime("%Y-%m-%d")
    conn = get_conn()
    rows = conn.execute(
        "SELECT chave FROM insights_ignorados WHERE ignorado_ate >= ?", (hoje,)
    ).fetchall()
    conn.close()
    return {"chaves": [r['chave'] for r in rows]}


# ── APELIDOS (nome fantasia) ──────────────────────────────────────────────

class NovoApelido(BaseModel):
    descricao_real: str            # nome como vem na fatura (após limpar_desc)
    apelido: str                   # nome fantasia que passa a aparecer
    aplicar_base: Optional[bool] = False  # troca todos os gastos com esse nome

@app.get("/apelidos")
def listar_apelidos():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM apelidos ORDER BY apelido").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/apelidos")
def salvar_apelido(body: NovoApelido):
    """Cadastra/atualiza um apelido. Com aplicar_base=True, troca em toda a base
    (lançamentos + staging) todos os gastos com esse nome e guarda o original."""
    real = (body.descricao_real or "").strip()
    apel = (body.apelido or "").strip()
    if not real or not apel:
        raise HTTPException(400, "Informe o nome real e o apelido.")
    conn = get_conn()
    conn.execute("""
        INSERT INTO apelidos (descricao_real, apelido) VALUES (?, ?)
        ON CONFLICT(descricao_real) DO UPDATE SET apelido=excluded.apelido
    """, (real, apel))
    atualizados = 0
    if body.aplicar_base:
        for tbl in ("lancamentos", "staging"):
            cur = conn.execute(f"""
                UPDATE {tbl}
                SET descricao_real = COALESCE(descricao_real, descricao),
                    descricao = ?
                WHERE COALESCE(descricao_real, descricao) = ?
            """, (apel, real))
            atualizados += cur.rowcount
    conn.commit(); conn.close()
    return {"ok": True, "atualizados": atualizados}

@app.delete("/apelidos/{descricao_real}")
def apagar_apelido(descricao_real: str):
    conn = get_conn()
    conn.execute("DELETE FROM apelidos WHERE descricao_real=?", (descricao_real,))
    conn.commit(); conn.close()
    return {"ok": True}


# ── METAS ─────────────────────────────────────────────────────────────────

_TIPOS_META = {'limite', 'reducao', 'superavit', 'acumulo'}
# Categorias de gasto permitidas em limite/reducao (exclui SA, I, F).
_CATS_GASTO_META = {'CA', 'S', 'L', 'C', 'T', 'M', 'E', 'A', 'B', 'R', 'O'}


@app.get("/metas")
def listar_metas():
    """Metas ativas da base atual (a base é resolvida pelo header X-Usuario)."""
    conn = get_conn()
    rows = conn.execute("SELECT * FROM metas WHERE ativa=1 ORDER BY id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/metas")
def criar_meta(body: NovaMeta):
    tipo = (body.tipo or '').strip().lower()
    if tipo not in _TIPOS_META:
        raise HTTPException(400, "Tipo de meta inválido.")
    if body.valor_alvo is None or body.valor_alvo <= 0:
        raise HTTPException(400, "O valor alvo deve ser maior que zero.")

    categoria = None
    reducao_modo = None
    acumulo_meses = None
    acumulo_mes_fim = None

    if tipo in ('limite', 'reducao'):
        categoria = (body.categoria or '').strip().upper()
        if categoria not in _CATS_GASTO_META:
            raise HTTPException(400, "Categoria inválida para esse tipo de meta.")

    if tipo == 'reducao':
        reducao_modo = (body.reducao_modo or '').strip().lower()
        if reducao_modo not in ('absoluto', 'percentual'):
            raise HTTPException(400, "reducao_modo deve ser 'absoluto' ou 'percentual'.")
        if reducao_modo == 'percentual' and body.valor_alvo > 100:
            raise HTTPException(400, "Redução percentual não pode passar de 100%.")

    if tipo == 'acumulo':
        acumulo_meses = body.acumulo_meses
        if not acumulo_meses or acumulo_meses < 2:
            raise HTTPException(400, "O acúmulo precisa de pelo menos 2 meses.")
        acumulo_mes_fim = (body.acumulo_mes_fim or '').strip()
        partes = acumulo_mes_fim.split('/')
        if len(partes) != 2 or not partes[0].isdigit() or not partes[1].isdigit() or len(partes[1]) != 4:
            raise HTTPException(400, "acumulo_mes_fim deve estar no formato M/AAAA.")

    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO metas (tipo, categoria, valor_alvo, reducao_modo, acumulo_meses, acumulo_mes_fim)
        VALUES (?,?,?,?,?,?)
    """, (tipo, categoria, body.valor_alvo, reducao_modo, acumulo_meses, acumulo_mes_fim))
    mid = cur.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM metas WHERE id=?", (mid,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/metas/{mid}")
def deletar_meta(mid: int):
    """Soft delete (ativa=0). A posse é garantida pela base (arquivo do usuário)."""
    conn = get_conn()
    conn.execute("UPDATE metas SET ativa=0 WHERE id=?", (mid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── VERIFICAÇÃO DE BASE ───────────────────────────────────────────────────

@app.get("/verificar")
def verificar_base():
    """Roda checagem de qualidade na base e retorna, por problema, as linhas
    afetadas (editáveis no pop-up Verificar base do frontend)."""
    conn = get_conn()
    problemas = []
    regras_usuario = _regras_usuario(conn)

    CATS_VALIDAS = {'SA','I','F','CA','S','E','A','T','M','C','B','R','L','O'}

    def _sugestao_para(it):
        """Passa o lançamento pelo mesmo pipeline dos gastos novos (regra → IA)
        e devolve uma categoria sugerida (pré-selecionável no frontend)."""
        desc = it['descricao'] or ''
        is_pix = (it['fonte'] or '').lower() == 'pix' or 'PIX' in desc.upper()
        cat, confianca, fonte = categorizar(
            desc, it['valor'] or 0, it['tipo'] or 'Débito', usar_ia=True,
            is_pix=is_pix, regras_usuario=regras_usuario
        )
        return {"categoria": cat, "confianca": confianca, "fonte": fonte}

    def add(tipo, descricao_fn, rows, sugerir=False):
        itens = [dict(r) for r in rows]
        if not itens:
            return
        if sugerir:
            # Roda regra+IA em paralelo (a IA é 1 requisição HTTP por item).
            with ThreadPoolExecutor(max_workers=8) as ex:
                for it, s in zip(itens, ex.map(_sugestao_para, itens)):
                    it["sugestao"] = s
        problemas.append({
            "tipo": tipo,
            "quantidade": len(itens),
            "descricao": descricao_fn(len(itens)),
            "itens": itens,
        })

    # 1. Sem categoria
    rows = conn.execute(
        "SELECT * FROM lancamentos WHERE categoria IS NULL OR categoria=''"
    ).fetchall()
    add("sem_categoria", lambda n: f"{n} lançamento(s) sem categoria", rows, sugerir=True)

    # 2. Categoria inválida
    rows = conn.execute(
        "SELECT * FROM lancamentos WHERE categoria IS NOT NULL AND categoria!=''"
    ).fetchall()
    invalidas = [r for r in rows if r['categoria'] not in CATS_VALIDAS]
    add("categoria_invalida", lambda n: f"{n} lançamento(s) com categoria inválida", invalidas, sugerir=True)

    # 3. Duplicatas exatas (todas as linhas que pertencem a um grupo duplicado)
    rows = conn.execute("""
        SELECT * FROM lancamentos l
        WHERE (SELECT COUNT(*) FROM lancamentos x
               WHERE x.data=l.data AND x.descricao=l.descricao AND x.valor=l.valor) > 1
        ORDER BY l.descricao, l.data
    """).fetchall()
    add("duplicatas", lambda n: f"{n} lançamento(s) duplicado(s)", rows)

    # 4. Datas inválidas
    rows = conn.execute(
        "SELECT * FROM lancamentos WHERE data NOT LIKE '__/__/____'"
    ).fetchall()
    add("data_invalida", lambda n: f"{n} lançamento(s) com data inválida", rows)

    # 5. Valores zerados (exceto F e I)
    rows = conn.execute(
        "SELECT * FROM lancamentos WHERE valor=0 AND categoria NOT IN ('F','I')"
    ).fetchall()
    add("valor_zero", lambda n: f"{n} lançamento(s) com valor zero", rows)

    # 6. Mês inconsistente com a data (só linhas com data válida — as inválidas já caem no item 4)
    rows = conn.execute("""
        SELECT * FROM lancamentos
        WHERE data LIKE '__/__/____'
          AND mes != (CAST(SUBSTR(data,4,2) AS INTEGER) || '/' || SUBSTR(data,7,4))
    """).fetchall()
    add("mes_inconsistente", lambda n: f"{n} lançamento(s) com mês inconsistente com a data", rows)

    conn.close()
    return {
        "ok": len(problemas) == 0,
        "total_problemas": len(problemas),
        "problemas": problemas
    }


# ── STATUS ────────────────────────────────────────────────────────────────

@app.get("/status")
def status_base():
    """
    Retorna sempre as datas mais recentes de débito e crédito
    direto da tabela lancamentos — inclui histórico carregado.
    """
    conn = get_conn()

    r_deb = conn.execute("""
        SELECT data as ultima FROM lancamentos
        WHERE tipo='Débito'
        ORDER BY SUBSTR(data,7,4) DESC,
                 SUBSTR(data,4,2) DESC,
                 SUBSTR(data,1,2) DESC
        LIMIT 1
    """).fetchone()

    r_cred = conn.execute("""
        SELECT data as ultima FROM lancamentos
        WHERE tipo='Crédito'
        ORDER BY SUBSTR(data,7,4) DESC,
                 SUBSTR(data,4,2) DESC,
                 SUBSTR(data,1,2) DESC
        LIMIT 1
    """).fetchone()

    conn.close()

    return {
        "ultimo_dado_debito":  r_deb['ultima']  if r_deb  else None,
        "ultimo_dado_credito": r_cred['ultima'] if r_cred else None,
    }

# ── LANÇAMENTOS ───────────────────────────────────────────────────────────

@app.get("/lancamentos")
def listar_lancamentos(mes: Optional[str]=None, revisado: Optional[int]=None, categoria: Optional[str]=None):
    conn = get_conn()
    q = "SELECT * FROM lancamentos WHERE 1=1"
    params = []
    if mes:       q += " AND mes=?";       params.append(mes)
    if revisado is not None: q += " AND revisado=?"; params.append(revisado)
    if categoria: q += " AND categoria=?"; params.append(categoria)
    # Ordenação cronológica real (data é texto dd/mm/yyyy): mais recentes primeiro
    q += (" ORDER BY SUBSTR(data,7,4) DESC,"
          " SUBSTR(data,4,2) DESC,"
          " SUBSTR(data,1,2) DESC")
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.patch("/lancamentos/{lid}")
def atualizar_lancamento(lid: int, body: AtualizarLancamento):
    conn = get_conn()
    fields, params = [], []
    if body.descricao is not None:
        atual = conn.execute("SELECT descricao, descricao_real FROM lancamentos WHERE id=?", (lid,)).fetchone()
        if atual is not None and not atual['descricao_real']:
            # Preserva o nome ORIGINAL da fatura na primeira renomeação.
            fields.append("descricao_real=?"); params.append(atual['descricao'])
        fields.append("descricao=?"); params.append(body.descricao)
    if body.categoria is not None: fields.append("categoria=?"); params.append(body.categoria)
    if body.viagem is not None:    fields.append("viagem=?");    params.append(body.viagem)
    if body.revisado is not None:  fields.append("revisado=?");  params.append(body.revisado)
    if body.data is not None:      fields.append("data=?");      params.append(body.data)
    if body.mes is not None:       fields.append("mes=?");       params.append(body.mes)
    if body.valor is not None:
        # Ao mudar o valor, recalcula crédito/débito para manter consistência.
        fields.append("valor=?");   params.append(body.valor)
        fields.append("credito=?"); params.append(body.valor if body.valor > 0 else None)
        fields.append("debito=?");  params.append(abs(body.valor) if body.valor < 0 else None)
    if not fields: raise HTTPException(400, "Nada para atualizar")
    params.append(lid)
    conn.execute(f"UPDATE lancamentos SET {','.join(fields)} WHERE id=?", params)
    conn.commit(); conn.close()
    return {"ok": True}

@app.delete("/lancamentos/{lid}")
def excluir_lancamento(lid: int):
    conn = get_conn()
    conn.execute("DELETE FROM lancamentos WHERE id=?", (lid,))
    conn.commit(); conn.close()
    return {"ok": True}


# ── VIAGENS ───────────────────────────────────────────────────────────────

@app.get("/viagens")
def listar_viagens():
    conn = get_conn()
    # Ordenação cronológica (data_inicio é texto dd/mm/yyyy): mais recentes primeiro
    viagens = conn.execute(
        "SELECT * FROM viagens ORDER BY"
        " SUBSTR(data_inicio,7,4) DESC,"
        " SUBSTR(data_inicio,4,2) DESC,"
        " SUBSTR(data_inicio,1,2) DESC"
    ).fetchall()
    result = []
    for v in viagens:
        lancamentos = conn.execute("SELECT * FROM lancamentos WHERE viagem=? ORDER BY data", (v['destino'],)).fetchall()
        gastos_cat = {}; total = 0
        for l in lancamentos:
            cat = l['categoria'] or 'O'
            if cat not in ('SA','I','F'):
                # Gasto líquido: débito (valor<0) entra positivo; crédito/estorno
                # (valor>0) ABATE o total. Antes usava abs() e o crédito somava.
                gasto = -(l['valor'] or 0)
                gastos_cat[cat] = gastos_cat.get(cat, 0) + gasto
                total += gasto
        result.append({**dict(v), 'total': round(total,2), 'por_categoria': {k: round(v2,2) for k,v2 in gastos_cat.items()}, 'num_lancamentos': len(lancamentos)})
    conn.close()
    return result

@app.post("/viagens")
def criar_viagem(body: NovaViagem):
    conn = get_conn()
    existe = conn.execute("SELECT id FROM viagens WHERE destino=?", (body.destino,)).fetchone()
    if existe:
        conn.execute(
            "UPDATE viagens SET data_inicio=?, data_fim=?, card=COALESCE(?, card) WHERE id=?",
            (body.data_inicio, body.data_fim, body.card, existe['id']))
    else:
        conn.execute(
            "INSERT INTO viagens (destino, data_inicio, data_fim, card) VALUES (?,?,?,?)",
            (body.destino, body.data_inicio, body.data_fim, body.card))
    conn.commit(); conn.close()
    return {"ok": True}

@app.delete("/viagens/{destino}")
def apagar_viagem(destino: str):
    """Apaga o registro/card da viagem e DESMARCA a viagem dos gastos."""
    conn = get_conn()
    conn.execute("DELETE FROM viagens WHERE destino=?", (destino,))
    conn.execute("UPDATE lancamentos SET viagem=NULL WHERE viagem=?", (destino,))
    conn.commit(); conn.close()
    return {"ok": True}

@app.get("/viagens/cards")
def listar_cards_viagem():
    """Lista os arquivos de imagem disponíveis como card (frontend/public/viagens)."""
    pasta = os.path.join(os.path.dirname(__file__), "..", "frontend", "public", "viagens")
    try:
        arquivos = sorted(
            f for f in os.listdir(pasta)
            if f.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))
        )
    except Exception:
        arquivos = []
    return {"cards": arquivos}


# ── DASHBOARD ─────────────────────────────────────────────────────────────

def _parse_mes(m):
    """'mes' é texto 'M/AAAA'. Retorna (ano, mes) ou None se malformado."""
    try:
        p = m.split('/')
        return (int(p[1]), int(p[0]))
    except Exception:
        return None


def _montar_metrica(mes, d):
    """Monta as métricas de um mês a partir do dict {categoria: soma_valor}."""
    receita = d.get('SA', 0)
    investimentos = d.get('I', 0)
    despesas = sum(v for k, v in d.items() if k not in ('SA', 'I', 'F') and v < 0)
    return {
        'mes': mes,
        'receita': round(receita, 2),
        'despesas': round(abs(despesas), 2),
        'investimentos': round(abs(investimentos), 2),
        'superavit': round(receita + despesas, 2),
        'categorias': {k: round(v, 2) for k, v in d.items()},
    }


@app.get("/dashboard")
def dashboard(meses: int = 6, ytd: bool = False, em_andamento: bool = False):
    """
    Métricas mensais agregadas.
    - meses: janela de meses mais recentes (ignorado se ytd=True).
    - ytd: só os meses do ano corrente.
    - em_andamento: inclui o mês corrente (incompleto) e adiciona o bloco
      'em_andamento' com a comparação do "período corrido" (dia 1 até hoje)
      contra o mesmo intervalo de dias do mês anterior.
    """
    conn = get_conn()
    rows = conn.execute("""
        SELECT mes, categoria, valor, data FROM lancamentos
        WHERE categoria != 'F' AND categoria IS NOT NULL ORDER BY mes
    """).fetchall()

    hoje = datetime.now()
    mes_corrente = (hoje.year, hoje.month)
    # Último mês completo = mês anterior ao atual
    if hoje.month == 1:
        limite = (hoje.year - 1, 12)
    else:
        limite = (hoje.year, hoje.month - 1)
    limite_efetivo = mes_corrente if em_andamento else limite

    def sort_mes(m):
        return _parse_mes(m) or (0, 0)

    por_mes = {}
    for r in rows:
        pm = _parse_mes(r['mes'])
        if not pm or pm > limite_efetivo:
            continue
        por_mes.setdefault(r['mes'], {})
        por_mes[r['mes']][r['categoria']] = por_mes[r['mes']].get(r['categoria'], 0) + r['valor']

    meses_ord = sorted(por_mes.keys(), key=sort_mes)
    if ytd:
        ultimos = [m for m in meses_ord if (_parse_mes(m) or (0, 0))[0] == hoje.year]
    else:
        ultimos = meses_ord[-meses:] if len(meses_ord) >= meses else meses_ord

    metricas = [_montar_metrica(m, por_mes[m]) for m in ultimos]

    ultimo = metricas[-1] if metricas else {}
    positivos = sum(1 for m in metricas if m['superavit'] > 0)
    indice = round((positivos / len(metricas)) * 100) if metricas else 0

    resp = {"indice_saude": indice, "metricas_mensais": metricas, "ultimo_mes": ultimo}

    # Bloco do "mês em andamento": comparação período corrido (dia 1..hoje)
    if em_andamento:
        mes_corrente_str = f"{hoje.month}/{hoje.year}"
        if hoje.month == 1:
            ano_ant, mes_ant = hoje.year - 1, 12
        else:
            ano_ant, mes_ant = hoje.year, hoje.month - 1
        mes_ant_str = f"{mes_ant}/{ano_ant}"
        corte = hoje.day

        def agregar_corrido(mes_str):
            d = {}
            for r in rows:
                if r['mes'] != mes_str:
                    continue
                try:
                    dia = int(r['data'][:2])
                except Exception:
                    continue
                if dia > corte:
                    continue
                d[r['categoria']] = d.get(r['categoria'], 0) + r['valor']
            return _montar_metrica(mes_str, d)

        resp['em_andamento'] = {
            'mes': mes_corrente_str,
            'mes_anterior': mes_ant_str,
            'dia_corte': corte,
            'corrente': agregar_corrido(mes_corrente_str),
            'anterior_corrido': agregar_corrido(mes_ant_str),
        }

    conn.close()
    return resp


# ── EXPORTAR ──────────────────────────────────────────────────────────────

@app.get("/exportar")
def exportar_historico():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import date as date_type

    conn = get_conn()
    rows = conn.execute("""
        SELECT mes, data, descricao, credito, debito, valor, categoria, tipo, viagem
        FROM lancamentos
        ORDER BY SUBSTR(data,7,4) DESC, SUBSTR(data,4,2) DESC, SUBSTR(data,1,2) DESC
    """).fetchall()
    conn.close()

    wb = Workbook(); ws = wb.active; ws.title = "Lançamentos"
    headers = ['Mês','Data','Descrição','Crédito(R$)','Débito(R$)','Valor','Categoria','Tipo','Viagem']
    hf = PatternFill('solid', start_color='1F4E79')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.alignment = Alignment(horizontal='center')
    for col, w in enumerate([10,12,45,14,14,14,12,10,20], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for ri, row in enumerate(rows, 2):
        try:
            p = row['data'].split('/')
            dv = date_type(int(p[2]), int(p[1]), int(p[0]))
        except: dv = row['data']
        vals = [row['mes'], dv, row['descricao'], row['credito'], row['debito'],
                row['valor'], row['categoria'], row['tipo'], row['viagem']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci == 2 and isinstance(dv, date_type): cell.number_format = 'DD/MM/YYYY'
            if ci in [4,5,6]: cell.number_format = '#,##0.00'
        if ri % 2 == 0:
            for ci in range(1,10): ws.cell(row=ri, column=ci).fill = PatternFill('solid', start_color='EBF3FB')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:I{len(rows)+1}"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    nome = f"historico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={nome}"})


# ── AVALIAÇÃO DO MODELO ───────────────────────────────────────────────────

def _amostra_aleatoria(rows, limite, seed):
    """Amostra aleatória de 'limite' linhas (ou todas, se limite for None/maior).

    Usa um seed para que a prévia e a execução selecionem exatamente as mesmas
    observações. Preserva a ordem original (para o progresso ficar estável).
    """
    if limite is None or limite >= len(rows):
        return list(rows)
    rng = random.Random(seed)
    indices = sorted(rng.sample(range(len(rows)), limite))
    return [rows[i] for i in indices]


@app.get("/avaliar/previa")
def avaliar_previa(limite_ia: Optional[int] = None, seed: Optional[int] = None):
    """Retorna quantos lançamentos vão chamar a IA antes de rodar a avaliação."""
    conn = get_conn()
    rows = conn.execute(
        "SELECT descricao, categoria, fonte FROM lancamentos "
        "WHERE categoria IS NOT NULL AND categoria != '' "
        "ORDER BY id"
    ).fetchall()
    regras_usuario = _regras_usuario(conn)
    conn.close()

    rows = [r for r in rows if r['categoria'] in CATEGORIAS_VALIDAS]
    total_geral = len(rows)
    rows_aval = _amostra_aleatoria(rows, limite_ia, seed)
    total = len(rows_aval)

    vai_ia = 0
    for row in rows_aval:
        is_pix = (row['fonte'] or '').lower() == 'pix' or 'PIX' in row['descricao'].upper()
        if not is_pix:
            _, status = categorizar_por_regra(row['descricao'], regras_usuario)
            if status == 'sem_regra':
                vai_ia += 1

    return {
        'total': total,
        'total_geral': total_geral,
        'vai_ia': vai_ia,
        'vai_ia_efetivo': vai_ia,
        'ia_disponivel': bool(ANTHROPIC_API_KEY),
    }


@app.get("/avaliar")
def avaliar_modelo(usar_ia: bool = False, limite_ia: Optional[int] = None, seed: Optional[int] = None):
    """
    Reclassifica todos os lançamentos e transmite progresso via SSE.
    Evento 'progress': {"type":"progress","done":N,"total":M}
    Evento 'result':   {"type":"result", ...métricas...}

    Com limite_ia definido, avalia uma amostra ALEATÓRIA de tamanho limite_ia
    (mesmo seed → mesma amostra que a prévia mostrou).
    """
    CATS_VALIDAS = CATEGORIAS_VALIDAS

    conn = get_conn()
    rows = conn.execute(
        "SELECT descricao, valor, tipo, categoria, fonte FROM lancamentos "
        "WHERE categoria IS NOT NULL AND categoria != '' "
        "ORDER BY id"
    ).fetchall()
    regras_usuario = _regras_usuario(conn)
    conn.close()

    def stream():
        total = 0
        corretos = 0
        por_categoria_real: dict = {}
        predicoes_por_cat: dict  = {}
        por_confianca = {
            'verde':    {'total': 0, 'correto': 0},
            'amarelo':  {'total': 0, 'correto': 0},
            'vermelho': {'total': 0, 'correto': 0},
        }
        por_fonte: dict = {}
        confusoes: dict = {}
        erros_lista = []
        ia_lista = []

        valid_rows = [r for r in rows if r['categoria'] in CATS_VALIDAS]
        if usar_ia and limite_ia is not None:
            valid_rows = _amostra_aleatoria(valid_rows, limite_ia, seed)
        n_total = len(valid_rows)

        for i, row in enumerate(valid_rows):
            cat_real = row['categoria']
            desc     = row['descricao']
            valor    = row['valor']
            tipo     = row['tipo']
            is_pix   = (row['fonte'] or '').lower() == 'pix'

            cat_pred, confianca, fonte_pred = categorizar(
                desc, valor, tipo, usar_ia=usar_ia, is_pix=is_pix, regras_usuario=regras_usuario
            )

            total += 1

            if cat_real not in por_categoria_real:
                por_categoria_real[cat_real] = {'total': 0, 'correto': 0}
            por_categoria_real[cat_real]['total'] += 1

            conf_key = confianca if confianca in por_confianca else 'vermelho'
            por_confianca[conf_key]['total'] += 1

            f = fonte_pred or 'sem_pred'
            if f not in por_fonte:
                por_fonte[f] = {'total': 0, 'correto': 0}
            por_fonte[f]['total'] += 1

            acertou = (cat_pred is not None) and (cat_pred == cat_real)

            if fonte_pred == 'ia':
                ia_lista.append({
                    'descricao': desc,
                    'valor':     valor,
                    'real':      cat_real,
                    'pred':      cat_pred,
                    'confianca': confianca,
                    'acertou':   acertou,
                })

            if acertou:
                corretos += 1
                por_categoria_real[cat_real]['correto'] += 1
                por_confianca[conf_key]['correto'] += 1
                por_fonte[f]['correto'] += 1
            elif cat_pred is not None and confianca != 'vermelho':
                par = f"{cat_real}→{cat_pred}"
                confusoes[par] = confusoes.get(par, 0) + 1
                erros_lista.append({
                    'descricao':  desc,
                    'real':       cat_real,
                    'pred':       cat_pred,
                    'confianca':  confianca,
                    'fonte_pred': fonte_pred,
                })

            if cat_pred is not None and confianca != 'vermelho':
                if cat_pred not in predicoes_por_cat:
                    predicoes_por_cat[cat_pred] = []
                predicoes_por_cat[cat_pred].append(cat_real)

            # Emite progresso a cada 10 registros ou no último
            if (i + 1) % 10 == 0 or (i + 1) == n_total:
                yield f"data: {json.dumps({'type':'progress','done':i+1,'total':n_total})}\n\n"

        # Calcula métricas finais
        stats_categoria = {}
        for cat, dados in por_categoria_real.items():
            t = dados['total']
            c = dados['correto']
            recall = c / t if t > 0 else 0
            preds = predicoes_por_cat.get(cat, [])
            tp = sum(1 for r in preds if r == cat)
            precisao = tp / len(preds) if preds else 0
            f1 = (2 * precisao * recall / (precisao + recall)) if (precisao + recall) > 0 else 0
            stats_categoria[cat] = {
                'total': t, 'correto': c,
                'recall': round(recall, 3),
                'precisao': round(precisao, 3),
                'f1': round(f1, 3),
            }

        precisa_ajuda  = por_confianca['vermelho']['total']
        preditos       = total - precisa_ajuda
        acuracia_total = round(corretos / total, 4) if total else 0
        acuracia_pred  = round(corretos / preditos, 4) if preditos else 0
        top_confusoes  = sorted(
            [{'par': k, 'count': v} for k, v in confusoes.items()],
            key=lambda x: -x['count']
        )[:12]

        resultado = {
            'type': 'result',
            'total': total, 'corretos': corretos,
            'precisa_ajuda': precisa_ajuda, 'preditos': preditos,
            'acuracia_total': acuracia_total, 'acuracia_preditos': acuracia_pred,
            'por_categoria': stats_categoria, 'por_confianca': por_confianca,
            'por_fonte': por_fonte, 'top_confusoes': top_confusoes,
            'erros_amostra': erros_lista[:30],
            'ia_lancamentos': ia_lista,
            'usar_ia': usar_ia, 'ia_disponivel': bool(ANTHROPIC_API_KEY),
        }
        yield f"data: {json.dumps(resultado)}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


# ── DIAGNÓSTICO ─────────────────────────────────────────────────────────────

@app.get("/diag")
def diag():
    """Estado do servidor vivo: confirma se a extração via IA está ativa.

    extraction_mode == 'ia'   → uploads usam a extração via IA (correta).
    extraction_mode == 'regex'→ sem chave; usa parser regex (best-effort).
    """
    import parser as _parser
    return {
        "api_key_present": bool(ANTHROPIC_API_KEY),       # de categorizer
        "parser_api_key_present": bool(_parser.ANTHROPIC_API_KEY),
        "pdftotext": _parser._achar_pdftotext() or None,
        "extraction_mode": "ia" if _parser.ANTHROPIC_API_KEY else "regex",
        "tem_extrair_lancamentos_ia": hasattr(_parser, "extrair_lancamentos_ia"),
    }


# ── USUÁRIOS / BASES ──────────────────────────────────────────────────────

class NovoUsuario(BaseModel):
    nome: str
    origem: Optional[str] = None   # base mestre p/ clonar; None = base vazia

@app.get("/usuarios")
def get_usuarios():
    """Lista as bases existentes, o usuário da requisição e o padrão."""
    return {"usuarios": listar_usuarios(), "atual": get_usuario(), "padrao": USUARIO_PADRAO}

@app.post("/usuarios")
def post_usuario(body: NovoUsuario):
    """Cria uma base nova: vazia, ou clonando uma base mestre (campo 'origem')."""
    if not body.nome.strip():
        raise HTTPException(400, "Informe um nome para a base.")
    nome = sanitizar_usuario(body.nome)
    if nome in set(listar_usuarios()):
        raise HTTPException(400, "Já existe uma base com esse nome.")
    if body.origem and sanitizar_usuario(body.origem) not in set(listar_usuarios()):
        raise HTTPException(400, "Base mestre de origem não encontrada.")
    u = criar_usuario(nome, origem=body.origem)
    _bases_inicializadas.add(u)
    return {"ok": True, "usuario": u}

@app.delete("/usuarios/{nome}")
def delete_usuario(nome: str):
    """Apaga a base (não permitido para a base padrão)."""
    try:
        apagar_usuario(nome)
    except ValueError as e:
        raise HTTPException(400, str(e))
    _bases_inicializadas.discard(sanitizar_usuario(nome))
    return {"ok": True}

@app.post("/usuarios/{nome}/zerar")
def post_zerar_usuario(nome: str):
    """Esvazia a base (mantém o schema) — para testar incorporação do zero."""
    zerar_usuario(nome)
    return {"ok": True}


# ── ROOT ──────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok", "app": "Finanças Pessoais"}
