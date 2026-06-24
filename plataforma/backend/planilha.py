"""Importação de base antiga via planilha Excel.

Lê um .xlsx, deixa o usuário apontar a linha do 1º dado e as colunas de
data/descrição/categoria/valor, e propõe um de-para entre as categorias da
planilha e as 14 categorias da plataforma. Os dados entram no staging (igual
ao PDF) — o usuário confere e incorpora pela aba Revisão.
"""

import io
import re
import unicodedata
from datetime import datetime, date

import openpyxl


# 14 categorias da plataforma (mesmos códigos do resto do sistema)
CATS_NOME = {
    'SA': 'Salário', 'I': 'Investimento', 'F': 'Fatura', 'CA': 'Casa',
    'S': 'Saúde', 'E': 'Estudo', 'A': 'Assinatura', 'T': 'Transporte',
    'M': 'Mercado', 'C': 'Comida', 'B': 'Bens', 'R': 'Roupa',
    'L': 'Lazer', 'O': 'Outros',
}

# Sinônimos/palavras-chave para o de-para heurístico (rótulo da planilha → código).
_SINONIMOS = {
    'SA': ['salario', 'renda', 'ordenado', 'provento', 'vencimento', 'receita', 'rendimentos do trabalho'],
    'I':  ['investimento', 'aplicacao', 'rendimento', 'poupanca', 'tesouro', 'acoes', 'aporte', 'dividendos'],
    'F':  ['fatura', 'cartao de credito', 'pagamento de fatura'],
    'CA': ['casa', 'moradia', 'aluguel', 'condominio', 'agua', 'luz', 'energia', 'gas', 'internet', 'telefone', 'iptu', 'lar', 'contas de casa', 'utilidades'],
    'S':  ['saude', 'farmacia', 'remedio', 'medico', 'hospital', 'exame', 'plano de saude', 'dentista', 'academia', 'suplemento'],
    'E':  ['estudo', 'educacao', 'curso', 'escola', 'faculdade', 'livro', 'material escolar', 'mensalidade'],
    'A':  ['assinatura', 'streaming', 'netflix', 'spotify', 'aplicativo', 'recorrente', 'servico recorrente'],
    'T':  ['transporte', 'uber', 'combustivel', 'gasolina', 'posto', 'onibus', 'metro', 'pedagio', 'estacionamento', 'passagem', 'corrida'],
    'M':  ['mercado', 'supermercado', 'hortifruti', 'feira', 'mercearia', 'compras de mercado'],
    'C':  ['comida', 'restaurante', 'lanche', 'alimentacao', 'delivery', 'ifood', 'refeicao', 'cafe', 'padaria'],
    'B':  ['bens', 'eletronico', 'eletronicos', 'produto', 'movel', 'utilidade', 'presente', 'tecnologia'],
    'R':  ['roupa', 'vestuario', 'calcado', 'moda', 'sapato', 'loja de roupa'],
    'L':  ['lazer', 'entretenimento', 'viagem', 'bar', 'balada', 'cinema', 'show', 'ingresso', 'evento', 'hotel', 'passeio', 'diversao', 'festa', 'role'],
    'O':  ['outros', 'diversos', 'geral', 'tarifa', 'imposto', 'taxa', 'saque', 'juros', 'multa'],
}


# ── HELPERS DE LEITURA/PARSE ──────────────────────────────────────────────

def _norm(s):
    """minúsculas, sem acento, só [a-z0-9 ]."""
    s = str(s or '').strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^a-z0-9 ]', ' ', s).strip()


def _cell_str(v):
    """Converte uma célula para texto amigável (datas → dd/mm/aaaa)."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, date):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _get(row, idx):
    """Acesso seguro a uma coluna da linha (tupla)."""
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _parse_data(v):
    """Normaliza a data da célula para 'dd/mm/aaaa' (ou None)."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s[:10], fmt).strftime('%d/%m/%Y')
        except ValueError:
            continue
    m = re.match(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})', s)
    if m:
        d, mo, y = m.groups()
        y = ('20' + y) if len(y) == 2 else y
        try:
            return datetime(int(y), int(mo), int(d)).strftime('%d/%m/%Y')
        except ValueError:
            return None
    return None


def _parse_valor(v):
    """Converte a célula de valor em float. Aceita número, '1.234,56', 'R$ -50,00'."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[^\d,.\-]', '', str(v).strip())
    if not s or s in ('-', '.', ','):
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


# ── DE-PARA DE CATEGORIAS (heurística) ────────────────────────────────────

def sugerir_categoria(label):
    """Mapeia um rótulo de categoria da planilha para um código da plataforma.

    Estratégia: nome igual ao da categoria → código direto → sinônimos.
    Retorna None quando não há palpite (o usuário escolhe manualmente).
    """
    n = _norm(label)
    if not n:
        return None

    # 1. Nome igual ao da categoria da plataforma (ex.: 'lazer' → L)
    for code, nome in CATS_NOME.items():
        if n == _norm(nome):
            return code

    # 2. A planilha já usa o nosso código (ex.: 'CA', 'L')
    up = str(label).strip().upper()
    if up in CATS_NOME:
        return up

    # 3. Sinônimos: match exato, palavra inteira, ou substring para termos longos
    palavras = set(n.split())
    melhor = None
    for code, syns in _SINONIMOS.items():
        for syn in syns:
            sn = _norm(syn)
            if not sn:
                continue
            hit = (sn == n) or (sn in palavras) or (len(sn) >= 5 and sn in n)
            if hit and (melhor is None or len(sn) > melhor[1]):
                melhor = (code, len(sn))
    return melhor[0] if melhor else None


# ── API DO MÓDULO ─────────────────────────────────────────────────────────

def ler_previa(file_bytes, max_linhas=25):
    """Lê as primeiras linhas da planilha (aba ativa) para montar a grade de mapeamento."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    linhas, n_col = [], 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_linhas:
            break
        vals = [_cell_str(c) for c in row]
        n_col = max(n_col, len(vals))
        linhas.append(vals)
    total = ws.max_row
    wb.close()
    linhas = [v + [''] * (n_col - len(v)) for v in linhas]  # normaliza largura
    return {'colunas': n_col, 'total_linhas': total, 'linhas': linhas}


def categorias_distintas(file_bytes, col_categoria, linha_inicio):
    """Lista os valores distintos da coluna de categoria + sugestão de de-para + contagem."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    contagem = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 < linha_inicio:  # i é 0-based; linha_inicio é 1-based (como no Excel)
            continue
        label = _cell_str(_get(row, col_categoria)).strip()
        if not label:
            continue
        contagem[label] = contagem.get(label, 0) + 1
    wb.close()
    return [
        {'valor': label, 'sugestao': sugerir_categoria(label), 'n': n}
        for label, n in sorted(contagem.items(), key=lambda x: -x[1])
    ]


def parse_planilha(file_bytes, mapeamento, de_para):
    """Lê a planilha aplicando o mapeamento e devolve lançamentos prontos p/ o staging.

    mapeamento: {linha_inicio (1-based), col_data, col_descricao, col_valor,
                 col_categoria (opcional), despesa_positiva (bool)}
    de_para:    {rótulo_da_planilha: código_da_plataforma}
    """
    col_data = mapeamento['col_data']
    col_desc = mapeamento['col_descricao']
    col_valor = mapeamento['col_valor']
    col_cat = mapeamento.get('col_categoria')
    linha_inicio = int(mapeamento.get('linha_inicio', 1))
    despesa_positiva = bool(mapeamento.get('despesa_positiva', False))

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    lancs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 < linha_inicio:
            continue

        data = _parse_data(_get(row, col_data))
        if not data:
            continue
        desc = _cell_str(_get(row, col_desc)).strip()
        if not desc:
            continue
        valor = _parse_valor(_get(row, col_valor))
        if valor is None:
            continue

        cat_label = _cell_str(_get(row, col_cat)).strip() if col_cat is not None else ''
        categoria = de_para.get(cat_label) if cat_label else None
        if categoria not in CATS_NOME:
            categoria = None

        # Sinal: se a planilha tem só valores positivos, tratamos como despesa
        # (negativo), exceto salário, que é receita (positivo).
        if despesa_positiva:
            valor = abs(valor) if categoria == 'SA' else -abs(valor)

        try:
            dt = datetime.strptime(data, '%d/%m/%Y')
        except ValueError:
            continue
        lancs.append({
            'mes': f"{dt.month}/{dt.year}",
            'data': data,
            'descricao': desc,
            'valor': valor,
            'categoria': categoria,
            'tipo': 'Crédito' if valor >= 0 else 'Débito',
        })

    wb.close()
    return lancs
